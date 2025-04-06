import os
import tempfile
import uuid
from collections import defaultdict
from datetime import datetime

import openpyxl
import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl_image_loader import SheetImageLoader


def po_pre_process(path, filename):
    wb = openpyxl.load_workbook(os.path.join(path, filename))

    po_data = defaultdict(list)
    po_meta = defaultdict(dict)
    po_image = defaultdict(None)

    for sheet in wb:
        id = sheet[1][0].value
        if id is None:
            continue
        loader = SheetImageLoader(sheet)
        po_data[id] = []

        for idx in range(3, 500):
            if sheet[idx][1].value == 'Trims':
                break

            placement = sheet[idx][1].value
            body = sheet[idx][2].value
            material = sheet[idx][3].value
            color = sheet[idx][4].value
            sc_body = sheet[idx][8].value
            color_type = sheet[idx][10].value
            material_id = material.split(' ')[0]

            po_data[id].append({
                "placement": placement,
                "body": body,
                "material": material,
                "material_id": material_id,
                "color": color,
                "sc_body": sc_body,
                "color_type": color_type
            })

            po_meta[id] = {
                "sc_body": sc_body,
            }

        for row in sheet.iter_rows(max_row=500, max_col=100):
            flag = False
            for cell in row:
                if loader.image_in(cell.coordinate):
                    image = loader.get(cell.coordinate)
                    image.save(os.path.join(path, f"{id}.png"))
                    po_image[id] = Image(os.path.join(path, f"{id}.png"))
                    break
            if flag:
                break
    return po_data, po_meta, po_image


def fabric_pre_process(path, filename):
    wb = openpyxl.load_workbook(os.path.join(path, filename), data_only=True)
    ws = wb.active

    material_mill_map = defaultdict(dict)
    material_mill_color_type_map = defaultdict(dict)

    for idx in range(2, 10000):
        fabric_id = ws[idx][2].value
        color_type = ws[idx][4].value
        mill = ws[idx][0].value
        fod = ws[idx][11].value
        cw = ws[idx][9].value
        lt = ws[idx][16].value

        if fabric_id is None:
            break

        material_mill_map[fabric_id] = mill
        material_mill_color_type_map[f"{fabric_id}:{color_type}"] = {
            'mill': mill,
            'fod': fod,
            'cw': cw,
            'lt': lt
        }

    max_fabric_color_type_map = {}
    for k in material_mill_color_type_map.keys():
        a = k.split(":")[0]
        if a not in max_fabric_color_type_map.keys():
            max_fabric_color_type_map[a] = 0
        max_fabric_color_type_map[a] += 1
    max_fabric_count = max(max_fabric_color_type_map.values())

    return material_mill_map, material_mill_color_type_map, max_fabric_count


def create_order(po_data, po_image, material_mill_map, output_path):
    order = Workbook()
    order_active = order.active

    order_active.row_dimensions[1].height = 40
    order_active.column_dimensions['A'].width = 40
    order_active.column_dimensions['B'].width = 40
    order_active.column_dimensions['C'].width = 40
    order_active.column_dimensions['D'].width = 40
    order_active.column_dimensions['E'].width = 40
    order_active.column_dimensions['F'].width = 40

    order_active.cell(row=1, column=1).value = 'STYLE#'
    order_active.cell(row=1, column=1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    order_active.cell(row=1, column=1).font = Font(color='FFFFFF', bold=True)
    order_active.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    order_active.cell(row=1, column=2).value = 'SKETCH'
    order_active.cell(row=1, column=2).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    order_active.cell(row=1, column=2).font = Font(color='FFFFFF', bold=True)
    order_active.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')

    order_active.cell(row=1, column=3).value = 'COLOR'
    order_active.cell(row=1, column=3).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    order_active.cell(row=1, column=3).font = Font(color='FFFFFF', bold=True)
    order_active.cell(row=1, column=3).alignment = Alignment(horizontal='center', vertical='center')

    order_active.cell(row=1, column=4).value = 'Placement'
    order_active.cell(row=1, column=4).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    order_active.cell(row=1, column=4).font = Font(color='FFFFFF', bold=True)
    order_active.cell(row=1, column=4).alignment = Alignment(horizontal='center', vertical='center')

    order_active.cell(row=1, column=5).value = 'Material'
    order_active.cell(row=1, column=5).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    order_active.cell(row=1, column=5).font = Font(color='FFFFFF', bold=True)
    order_active.cell(row=1, column=5).alignment = Alignment(horizontal='center', vertical='center')

    order_active.cell(row=1, column=6).value = 'Mill'
    order_active.cell(row=1, column=6).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    order_active.cell(row=1, column=6).font = Font(color='FFFFFF', bold=True)
    order_active.cell(row=1, column=6).alignment = Alignment(horizontal='center', vertical='center')

    start_idx = 2

    for p_id, data in po_data.items():
        order_active.cell(row=start_idx, column=1).value = p_id
        order_active.cell(row=start_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
        order_active.add_image(po_image[p_id], f'B{start_idx}')

        for d in data:
            order_active.row_dimensions[start_idx].height = 220
            order_active.cell(row=start_idx, column=3).value = d['color']
            order_active.cell(row=start_idx, column=3).alignment = Alignment(horizontal='center', vertical='center')

            order_active.cell(row=start_idx, column=4).value = d['placement']
            order_active.cell(row=start_idx, column=4).alignment = Alignment(horizontal='center', vertical='center')

            order_active.cell(row=start_idx, column=5).value = d['material']
            order_active.cell(row=start_idx, column=5).alignment = Alignment(horizontal='center', vertical='center')

            m_id = "매핑불가능"
            if d['material_id'] in material_mill_map.keys():
                m_id = material_mill_map[d['material_id']]

            order_active.cell(row=start_idx, column=6).value = m_id
            order_active.cell(row=start_idx, column=6).alignment = Alignment(horizontal='center', vertical='center')
            start_idx += 1

    order.save(output_path)


def create_cost(po_data, po_image, po_meta, material_mill_map, material_mill_color_type_map, max_fabric_count,
                output_path):
    cost = Workbook()
    cost_active = cost.active

    cost_active.row_dimensions[1].height = 40
    cost_active.column_dimensions['A'].width = 40
    cost_active.column_dimensions['B'].width = 40
    cost_active.column_dimensions['C'].width = 40

    cost_active.cell(row=1, column=1).value = 'Image'
    cost_active.cell(row=1, column=1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    cost_active.cell(row=1, column=1).font = Font(color='FFFFFF', bold=True)
    cost_active.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    cost_active.cell(row=1, column=2).value = 'Style ID'
    cost_active.cell(row=1, column=2).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    cost_active.cell(row=1, column=2).font = Font(color='FFFFFF', bold=True)
    cost_active.cell(row=1, column=2).alignment = Alignment(horizontal='center', vertical='center')

    cost_active.cell(row=1, column=3).value = 'Component detail'
    cost_active.cell(row=1, column=3).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    cost_active.cell(row=1, column=3).font = Font(color='FFFFFF', bold=True)
    cost_active.cell(row=1, column=3).alignment = Alignment(horizontal='center', vertical='center')

    for idx in range(0, max_fabric_count):
        plus = idx * 7
        cost_active.cell(row=1, column=4 + plus).value = f"Fab{idx + 1} Mil"
        cost_active.cell(row=1, column=4 + plus).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                    fill_type='solid')
        cost_active.cell(row=1, column=4 + plus).font = Font(color='FFFFFF', bold=True)
        cost_active.cell(row=1, column=4 + plus).alignment = Alignment(horizontal='center', vertical='center')
        cost_active.column_dimensions[get_column_letter(4 + plus)].width = 40

        cost_active.cell(row=1, column=5 + plus).value = f"Fab{idx + 1} LT"
        cost_active.cell(row=1, column=5 + plus).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                    fill_type='solid')
        cost_active.cell(row=1, column=5 + plus).font = Font(color='FFFFFF', bold=True)
        cost_active.cell(row=1, column=5 + plus).alignment = Alignment(horizontal='center', vertical='center')
        cost_active.column_dimensions[get_column_letter(5 + plus)].width = 40

        cost_active.cell(row=1, column=6 + plus).value = f"Fab{idx + 1} CW"
        cost_active.cell(row=1, column=6 + plus).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                    fill_type='solid')
        cost_active.cell(row=1, column=6 + plus).font = Font(color='FFFFFF', bold=True)
        cost_active.cell(row=1, column=6 + plus).alignment = Alignment(horizontal='center', vertical='center')
        cost_active.column_dimensions[get_column_letter(6 + plus)].width = 40

        cost_active.cell(row=1, column=7 + plus).value = f"Fab{idx + 1} Part"
        cost_active.cell(row=1, column=7 + plus).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                    fill_type='solid')
        cost_active.cell(row=1, column=7 + plus).font = Font(color='FFFFFF', bold=True)
        cost_active.cell(row=1, column=7 + plus).alignment = Alignment(horizontal='center', vertical='center')
        cost_active.column_dimensions[get_column_letter(7 + plus)].width = 40

        cost_active.cell(row=1, column=8 + plus).value = f"Fab{idx + 1} Fabrication"
        cost_active.cell(row=1, column=8 + plus).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                    fill_type='solid')
        cost_active.cell(row=1, column=8 + plus).font = Font(color='FFFFFF', bold=True)
        cost_active.cell(row=1, column=8 + plus).alignment = Alignment(horizontal='center', vertical='center')
        cost_active.column_dimensions[get_column_letter(8 + plus)].width = 40

        cost_active.cell(row=1, column=9 + plus).value = f"Fab{idx + 1} Color"
        cost_active.cell(row=1, column=9 + plus).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                    fill_type='solid')
        cost_active.cell(row=1, column=9 + plus).font = Font(color='FFFFFF', bold=True)
        cost_active.cell(row=1, column=9 + plus).alignment = Alignment(horizontal='center', vertical='center')
        cost_active.column_dimensions[get_column_letter(9 + plus)].width = 40

        cost_active.cell(row=1, column=10 + plus).value = f"Fab{idx + 1} FOB / yds"
        cost_active.cell(row=1, column=10 + plus).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                     fill_type='solid')
        cost_active.cell(row=1, column=10 + plus).font = Font(color='FFFFFF', bold=True)
        cost_active.cell(row=1, column=10 + plus).alignment = Alignment(horizontal='center', vertical='center')
        cost_active.column_dimensions[get_column_letter(10 + plus)].width = 40

    start_idx = 2

    for p_id, data in po_data.items():
        cost_active.row_dimensions[start_idx].height = 220
        cost_active.add_image(po_image[p_id], f'A{start_idx}')
        cost_active.cell(row=start_idx, column=2).value = p_id
        cost_active.cell(row=start_idx, column=2).alignment = Alignment(horizontal='center', vertical='center')

        cost_active.cell(row=start_idx, column=3).value = po_meta[p_id]['sc_body']
        cost_active.cell(row=start_idx, column=3).alignment = Alignment(horizontal='center', vertical='center')

        plus = 0
        for d in data:
            m_id = '매핑불가능'
            if d['material_id'] in material_mill_map.keys():
                m_id = material_mill_map[d['material_id']]
            color_type_map = material_mill_color_type_map[f"{d['material_id']}:{d['color_type']}"]
            if len(color_type_map.keys()) == 0:
                color_type_map = {
                    'mill': m_id,
                    'cw': '매핑불가능',
                    'fod': '매핑불가능',
                    'lt': '매핑불가능',
                }

            cost_active.cell(row=start_idx, column=plus + 4).value = color_type_map['mill']
            cost_active.cell(row=start_idx, column=plus + 4).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

            cost_active.cell(row=start_idx, column=plus + 5).value = color_type_map['lt']
            cost_active.cell(row=start_idx, column=plus + 5).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

            cost_active.cell(row=start_idx, column=plus + 6).value = color_type_map['cw']
            cost_active.cell(row=start_idx, column=plus + 6).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

            cost_active.cell(row=start_idx, column=plus + 7).value = d['placement']
            cost_active.cell(row=start_idx, column=plus + 7).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

            cost_active.cell(row=start_idx, column=plus + 8).value = d['material']
            cost_active.cell(row=start_idx, column=plus + 8).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

            cost_active.cell(row=start_idx, column=plus + 9).value = d['color']
            cost_active.cell(row=start_idx, column=plus + 9).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')

            cost_active.cell(row=start_idx, column=plus + 10).value = color_type_map['fod']
            cost_active.cell(row=start_idx, column=plus + 10).alignment = Alignment(horizontal='center',
                                                                                    vertical='center')
            plus += 7

        start_idx += 1

    cost.save(output_path)


st.title("발주서와 코스팅")

if st.button("다시 하기"):
    st.rerun()

with st.form(key='form', clear_on_submit=True):
    po_uploaded_file = st.file_uploader("PO 차트를 업로드하세요", type=["xlsx"], key="po_uploaded_file")
    fabric_uploaded_file = st.file_uploader("Fabric mill 차트를 업로드하세요", type=["xlsx"], key="fabric_uploaded_file")
    submitted = st.form_submit_button("submit")

if po_uploaded_file and fabric_uploaded_file:
    with tempfile.TemporaryDirectory() as tmpdir:
        po_input_path = os.path.join(tmpdir, 'po.xlsx')
        with open(po_input_path, "wb") as f:
            f.write(po_uploaded_file.read())
        fabric_input_path = os.path.join(tmpdir, 'fabric.xlsx')
        with open(fabric_input_path, "wb") as f:
            f.write(fabric_uploaded_file.read())
        st.success("파일 업로드 완료! 데이터 처리 중 입니다")
        po_data, po_meta, po_image = po_pre_process(tmpdir, 'po.xlsx')
        material_mill_map, material_mill_color_type_map, max_fabric_count = fabric_pre_process(tmpdir, 'fabric.xlsx')

        order_file_name = f"order {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.xlsx"
        order_path = os.path.join(tmpdir, order_file_name)
        cost_file_name = f"cost {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.xlsx"
        cost_path = os.path.join(tmpdir, cost_file_name)
        create_order(po_data=po_data, po_image=po_image, material_mill_map=material_mill_map,
                     output_path=order_path)
        create_cost(po_data=po_data, po_meta=po_meta, po_image=po_image, material_mill_map=material_mill_map,
                    material_mill_color_type_map=material_mill_color_type_map, max_fabric_count=max_fabric_count,
                    output_path=cost_path)
        st.success("데이터 처리 완료! 다운로드 버튼을 눌러주세요!")
        st.download_button(label='발주서', data=open(order_path, "rb"), file_name=os.path.basename(order_path))
        st.download_button(label='코스팅', data=open(cost_path, "rb"), file_name=os.path.basename(cost_path))
